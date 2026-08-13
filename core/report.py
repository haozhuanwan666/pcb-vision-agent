"""
报表生成 Report
===============
Excel 质检报表（带样式）+ JSON 日志/报表落盘。
由项目早期 web_app.py 的报表逻辑重构而来。
"""
import json
from datetime import datetime
from typing import Dict, List, Optional

from ..utils.config import OUTPUT_LOGS, OUTPUT_REPORTS
from ..utils.file_utils import ensure_dir, timestamp_str


def generate_excel_report(
    results: List[Dict], output_path: Optional[str] = None
) -> Optional[str]:
    """
    生成 Excel 质检报表（带样式）
    results: [{filename, quality_level, total_defects, defect_by_class, verdict, ...}]
    返回报表路径；openpyxl 缺失时返回 None。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
    except ImportError:
        print("⚠️ openpyxl未安装，运行: pip install openpyxl")
        return None

    if output_path is None:
        output_path = str(OUTPUT_REPORTS / f"质检报告_{timestamp_str()}.xlsx")
    ensure_dir(OUTPUT_REPORTS)

    wb = Workbook()
    ws = wb.active
    ws.title = "质检报告"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = "PCB板缺陷检测质检报告"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center_align

    # 统计信息
    pass_count = sum(1 for r in results if r["total_defects"] == 0)
    fail_count = len(results) - pass_count
    pass_rate = pass_count / len(results) * 100 if results else 0
    ws["A3"] = "检测时间"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "检测总数"
    ws["B4"] = len(results)
    ws["A5"] = "合格数"
    ws["B5"] = pass_count
    ws["A6"] = "不合格数"
    ws["B6"] = fail_count
    ws["A7"] = "良率"
    ws["B7"] = f"{pass_rate:.2f}%"

    # 明细表头
    headers = ["序号", "文件名", "质量等级", "缺陷数量", "缺陷类型", "判定结果", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 明细数据
    for idx, r in enumerate(results, 1):
        row = 9 + idx
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=r.get("filename", "")).border = thin_border
        ws.cell(row=row, column=3, value=r.get("quality_level", "")).border = thin_border
        ws.cell(row=row, column=4, value=r.get("total_defects", 0)).border = thin_border
        defect_types = ", ".join(r.get("defect_by_class", {}).keys()) or "无"
        ws.cell(row=row, column=5, value=defect_types).border = thin_border
        verdict_cell = ws.cell(row=row, column=6, value=r.get("verdict", ""))
        verdict_cell.border = thin_border
        verdict_cell.alignment = center_align
        verdict_cell.fill = pass_fill if r.get("verdict") == "合格" else fail_fill
        ws.cell(row=row, column=7, value="").border = thin_border

    # 调整列宽
    widths = {"A": 8, "B": 30, "C": 12, "D": 12, "E": 25, "F": 12, "G": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    return output_path


def save_json_log(data, name_prefix: str = "batch") -> str:
    """保存 JSON 日志到 output/logs"""
    ensure_dir(OUTPUT_LOGS)
    path = str(OUTPUT_LOGS / f"{name_prefix}_{timestamp_str()}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def save_json_report(data, name_prefix: str = "检测报告") -> str:
    """保存 JSON 报表到 output/reports"""
    ensure_dir(OUTPUT_REPORTS)
    path = str(OUTPUT_REPORTS / f"{name_prefix}_{timestamp_str()}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path
